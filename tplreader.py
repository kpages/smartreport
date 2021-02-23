""" 
从Word 模板中识别标记信息 并按规则生成Excel 配置文件 
{{ var }}  : 条目配置的配置项
{{ table_{sheetname} }} : 以sheetname 为名称的表格
{{ line_table_{sheetname}_{x1:x2}_{y1,y2,y3} }} 使用 以sheetname 为名称的表格中的表格数据画图
line_{sheetname}_D1:D8_E,F,H
{
    sheetname: "sheet1",
    x: "D1:D8",
    y: "E,F,H"
}
"""
import os
import re
from docxtpl import DocxTemplate 
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl import Workbook, load_workbook

partter = r'{{([^\s]+)}}'

def read_config(text):
    # return type, config
    ms = re.findall(partter, text)
    if not ms:
        return None, None

    m = ms[0]
    if m.startswith("table_"):
        return "table", m[6:]
    elif m.startswith("line_"):
        secs = m.split("_")
        return "line", dict(sheetname=secs[1], x=secs[2], y=secs[3])
    else:
        return "item", m

def tpl_reader(word_path):
    tpl = DocxTemplate(word_path)
    config = {
        "items": [],
        "tables": [],
        "lines": []
    }

    for pg in tpl.docx.paragraphs:
        ctype, conf = read_config(pg.text)
        if ctype:
            config[f"{ctype}s"].append(conf)
        
    return config


def tpl2excel(word_path):
    config = tpl_reader(word_path)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "条目配置"
    for row, conf in enumerate(config.get("items", [])):
        ws1.cell(column=1, row=row+1, value=conf)
    
    for table_conf in config.get("tables", []):
        _ws = wb.create_sheet(title=table_conf)

    wb.save(filename = f"{word_path}.xlsx")


def get_item_config(ws):
    config = {}
    for x in range(1, ws.max_row+1):
        cell_key = ws.cell(row=x, column=1, value=None)
        cell_value = ws.cell(row=x, column=2, value=None)
        if cell_key.value:
            config[cell_key.value] = cell_value.value
    
    return config


def excel2word(excel_path, word_path, output_path):
    # 选择Excel 配置文件和Word 模板文件, 生成Word 报告
    tpl = DocxTemplate(word_path)
    wb = load_workbook(filename=excel_path)

    context = {}
    for sheetname in wb.sheetnames:
        ws1 = wb[sheetname]
        if sheetname == "条目配置":
            context.update(get_item_config(ws1))
        else:
            document = tpl.new_subdoc()
            context[f"table_{sheetname}"] = document
            table = document.add_table(rows=ws1.max_row, cols=ws1.max_column)
            table.style = "table"
            for x in range(1, ws1.max_row+1):
                for y in range(1, ws1.max_column+1):
                    c = ws1.cell(row=x, column=y, value=None)
                    if not type(c) == MergedCell:
                        table.rows[x-1].cells[y-1].text = str(c.value)

            for cell in ws1.merged_cells:
                # print(type(cell))
                cells = list(cell.cells)
                c_f = table.rows[cells[0][0]-1].cells[cells[0][1]-1]
                c_l = table.rows[cells[-1][0]-1].cells[cells[-1][1]-1]
                c_f.merge(c_l)

    tpl.render(context)
    tpl.save(output_path)


def get_tpls(output_path):
    files = os.listdir(output_path)
    return filter(lambda x: os.path.splitext(x)[1] == ".docx", files)
        



